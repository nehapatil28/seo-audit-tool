"""
export.py — SEOBit Full Report Exporter
========================================
Generates PDF and Excel reports from audit data.
"""

import io
import datetime

# ─────────────────────────────────────────────────────────────────────────────
#  HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def _now():
    return datetime.datetime.now().strftime("%Y-%m-%d %H:%M")

def _score_label(score):
    if score is None: return "N/A"
    if score >= 80: return "Good"
    if score >= 50: return "Needs Work"
    return "Poor"

def _ms_to_s(ms):
    if ms is None: return "N/A"
    return f"{round(ms/1000, 2)}s"

# ─────────────────────────────────────────────────────────────────────────────
#  PDF EXPORT
# ─────────────────────────────────────────────────────────────────────────────

def generate_pdf(data: dict) -> bytes:
    import math
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
        HRFlowable, KeepTogether, PageBreak, Flowable
    )
    from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
    from reportlab.pdfgen import canvas as rl_canvas

    # ══════════════════════════════════════════════
    #  COLOR PALETTE
    # ══════════════════════════════════════════════
    ORANGE      = colors.HexColor("#f06400")
    ORANGE_DARK = colors.HexColor("#c04e00")
    ORANGE_LITE = colors.HexColor("#fff3eb")
    DARK        = colors.HexColor("#1a1916")
    DARK2       = colors.HexColor("#2d2b28")
    GRAY        = colors.HexColor("#6b6963")
    GRAY_LITE   = colors.HexColor("#f5f4f1")
    BORDER      = colors.HexColor("#e2e0da")
    GREEN       = colors.HexColor("#16a34a")
    GREEN_BG    = colors.HexColor("#f0fdf4")
    RED         = colors.HexColor("#dc2626")
    RED_BG      = colors.HexColor("#fef2f2")
    AMBER       = colors.HexColor("#d97706")
    AMBER_BG    = colors.HexColor("#fffbeb")
    BLUE        = colors.HexColor("#2563eb")
    WHITE       = colors.white
    BLACK       = colors.black

    PW, PH = A4
    LM = RM = 18 * mm
    TM = BM = 16 * mm
    W  = PW - LM - RM   # usable content width

    # ══════════════════════════════════════════════
    #  SCORE COLOR HELPER
    # ══════════════════════════════════════════════
    def score_color(score):
        if score is None: return GRAY
        if score >= 80:   return GREEN
        if score >= 50:   return AMBER
        return RED

    def score_bg(score):
        if score is None: return GRAY_LITE
        if score >= 80:   return GREEN_BG
        if score >= 50:   return AMBER_BG
        return RED_BG

    def score_label(score):
        if score is None: return "N/A"
        if score >= 80:   return "Good"
        if score >= 50:   return "Needs Work"
        return "Poor"

    def rating_color(rating):
        r = (rating or "").lower()
        if r == "good":              return GREEN
        if r in ("poor", "fail"):    return RED
        if r == "needs-improvement": return AMBER
        return GRAY

    # ══════════════════════════════════════════════
    #  CUSTOM FLOWABLES
    # ══════════════════════════════════════════════
    class ScoreGauge(Flowable):
        """Circular arc gauge showing a score 0-100."""
        def __init__(self, score, label, size=52):
            super().__init__()
            self.raw_score = score
            self.score = score if score is not None else 0
            self.label = label
            self.size  = size
            self.width  = size
            self.height = size + 24

        def draw(self):
            c = self.canv
            cx, cy = self.size / 2, self.size / 2 + 18
            r = self.size / 2 - 4
            score = self.score

            # Background arc (full circle track)
            c.setStrokeColor(BORDER)
            c.setLineWidth(5)
            c.arc(cx - r, cy - r, cx + r, cy + r, startAng=0, extent=360)

            # Colored arc  (starts at top = 90°, goes clockwise as score increases)
            extent = -(score / 100) * 360
            c.setStrokeColor(score_color(self.raw_score))
            c.setLineWidth(5)
            if score > 0:
                c.arc(cx - r, cy - r, cx + r, cy + r, startAng=90, extent=extent)

            # Score text in center
            c.setFillColor(score_color(self.raw_score))
            c.setFont("Helvetica-Bold", 13)
            center_text = "N/A" if self.raw_score is None else str(int(round(score)))
            c.drawCentredString(cx, cy - 5, center_text)

            # Label below (with a clear gap below the ring)
            c.setFillColor(GRAY)
            c.setFont("Helvetica", 7)
            # wrap label
            words = self.label.split()
            line1 = " ".join(words[:2])
            line2 = " ".join(words[2:]) if len(words) > 2 else ""
            if line2:
                c.drawCentredString(cx, 9, line1)
                c.drawCentredString(cx, 1, line2)
            else:
                c.drawCentredString(cx, 5, line1)

    class SectionHeader(Flowable):
        """Full-width colored section banner."""
        def __init__(self, title, width, icon=""):
            super().__init__()
            self.title  = title
            self.icon   = icon
            self.width  = width
            self.height = 24

        def draw(self):
            c = self.canv
            c.setFillColor(DARK2)
            c.roundRect(0, 0, self.width, self.height, 4, fill=1, stroke=0)
            c.setFillColor(WHITE)
            c.setFont("Helvetica-Bold", 11)
            text = f"{self.icon}  {self.title}" if self.icon else self.title
            c.drawString(10, 7, text)

    class StatusBadge(Flowable):
        """Colored pill badge: Pass / Fail / Warn."""
        def __init__(self, status_text, w=54, h=14):
            super().__init__()
            self.status = status_text
            self.width  = w
            self.height = h

        def draw(self):
            c = self.canv
            s = self.status.lower()
            if any(x in s for x in ["yes","found","present","good","pass","ok","no noindex"]):
                bg, fg = GREEN, WHITE
            elif any(x in s for x in ["no","missing","fail","error","broken","poor"]):
                bg, fg = RED, WHITE
            elif any(x in s for x in ["warn","needs","improvement","partial"]):
                bg, fg = AMBER, WHITE
            else:
                bg, fg = GRAY_LITE, DARK
            c.setFillColor(bg)
            c.roundRect(0, 0, self.width, self.height, 6, fill=1, stroke=0)
            c.setFillColor(fg)
            c.setFont("Helvetica-Bold", 7)
            c.drawCentredString(self.width / 2, 4, self.status[:18])

    class ScoreBar(Flowable):
        """Horizontal bar showing score."""
        def __init__(self, score, bar_width=120, height=12):
            super().__init__()
            self.score = score if score is not None else 0
            self.width  = bar_width
            self.height = height

        def draw(self):
            c = self.canv
            # background track
            c.setFillColor(BORDER)
            c.roundRect(0, 2, self.width, self.height - 4, 4, fill=1, stroke=0)
            # filled portion
            filled = max(4, int((self.score / 100) * self.width))
            c.setFillColor(score_color(self.score))
            c.roundRect(0, 2, filled, self.height - 4, 4, fill=1, stroke=0)

    # ══════════════════════════════════════════════
    #  STYLES
    # ══════════════════════════════════════════════
    base = getSampleStyleSheet()

    def sty(name, **kw):
        # Auto-compute a sane leading (line height) from fontSize when not
        # explicitly provided — otherwise large fonts inherit the tiny
        # default leading from the "Normal" style and overlap adjacent text.
        if "leading" not in kw:
            fs = kw.get("fontSize", base["Normal"].fontSize)
            kw["leading"] = fs * 1.2
        return ParagraphStyle(name, parent=base["Normal"], **kw)

    COVER_BRAND = sty("cb",  fontSize=32, fontName="Helvetica-Bold", textColor=ORANGE, spaceAfter=0)
    COVER_TAG   = sty("ct",  fontSize=12, fontName="Helvetica",      textColor=GRAY,   spaceAfter=0)
    COVER_URL   = sty("cu",  fontSize=10, fontName="Helvetica-Bold", textColor=DARK2,  spaceAfter=0)
    COVER_DATE  = sty("cd",  fontSize=9,  fontName="Helvetica",      textColor=GRAY)

    H2    = sty("h2", fontSize=11, fontName="Helvetica-Bold", textColor=DARK,  spaceBefore=12, spaceAfter=5)
    H3    = sty("h3", fontSize=9,  fontName="Helvetica-Bold", textColor=DARK2, spaceBefore=8,  spaceAfter=3)
    BODY  = sty("bd", fontSize=9,  fontName="Helvetica",      textColor=DARK,  spaceAfter=3, leading=13)
    BODY_GRAY = sty("bg", fontSize=8, fontName="Helvetica",   textColor=GRAY,  spaceAfter=2)
    SMALL = sty("sm", fontSize=7.5, fontName="Helvetica",     textColor=GRAY)
    URL_S = sty("us", fontSize=8.5, fontName="Helvetica",     textColor=GRAY,  spaceAfter=2)

    # Cell styles for tables
    def cell(text, bold=False, color=DARK, size=8.5, align=TA_LEFT):
        fs = size; fn = "Helvetica-Bold" if bold else "Helvetica"
        return Paragraph(str(text), sty(f"_c{id(text)}", fontSize=fs,
                         fontName=fn, textColor=color,
                         alignment=align, leading=fs+3))

    # ══════════════════════════════════════════════
    #  TABLE FACTORIES
    # ══════════════════════════════════════════════
    def make_table(rows, col_ratios, header_row=True, compact=False):
        col_widths = [W * r for r in col_ratios]
        pad = 4 if compact else 6
        t = Table(rows, colWidths=col_widths, repeatRows=1 if header_row else 0)
        ts = [
            ("FONTSIZE",       (0,0), (-1,-1), 8.5),
            ("FONTNAME",       (0,0), (-1,-1), "Helvetica"),
            ("ROWBACKGROUNDS", (0,1), (-1,-1), [WHITE, GRAY_LITE]),
            ("LINEBELOW",      (0,0), (-1,-1), 0.3, BORDER),
            ("VALIGN",         (0,0), (-1,-1), "MIDDLE"),
            ("TOPPADDING",     (0,0), (-1,-1), pad),
            ("BOTTOMPADDING",  (0,0), (-1,-1), pad),
            ("LEFTPADDING",    (0,0), (-1,-1), 8),
            ("RIGHTPADDING",   (0,0), (-1,-1), 8),
        ]
        if header_row:
            ts += [
                ("BACKGROUND",  (0,0), (-1,0), ORANGE),
                ("TEXTCOLOR",   (0,0), (-1,0), WHITE),
                ("FONTNAME",    (0,0), (-1,0), "Helvetica-Bold"),
                ("FONTSIZE",    (0,0), (-1,0), 8.5),
                ("TOPPADDING",  (0,0), (-1,0), 7),
                ("BOTTOMPADDING",(0,0),(-1,0), 7),
                ("LINEABOVE",   (0,0), (-1,0), 0, WHITE),
            ]
        t.setStyle(TableStyle(ts))
        return t

    def two_col_kv(rows_data, header=None):
        """Simple 2-col key-value table."""
        rows = []
        if header:
            rows.append([cell(header[0], bold=True), cell(header[1], bold=True)])
        for k, v in rows_data:
            rows.append([cell(k), cell(str(v))])
        return make_table(rows, [0.55, 0.45], header_row=bool(header))

    def status_row(check, status_str, detail=""):
        """Row with colored status text."""
        s = status_str.lower()
        if any(x in s for x in ["yes","found","present","good","no noindex"]):
            sc = GREEN
        elif any(x in s for x in ["no","missing","fail","broken","poor"]):
            sc = RED
        elif any(x in s for x in ["warn","needs","partial"]):
            sc = AMBER
        else:
            sc = DARK
        return [
            cell(check),
            cell(status_str, color=sc, bold=True),
            cell(detail, color=GRAY, size=8),
        ]

    def hr(thick=0.5, space_before=6, space_after=8):
        return HRFlowable(width="100%", thickness=thick,
                          color=BORDER, spaceBefore=space_before, spaceAfter=space_after)

    # ══════════════════════════════════════════════
    #  PAGE TEMPLATE  (header/footer on each page)
    # ══════════════════════════════════════════════
    def on_page(canv, doc, url="", generated=""):
        canv.saveState()
        # Top bar
        canv.setFillColor(DARK2)
        canv.rect(0, PH - 10*mm, PW, 10*mm, fill=1, stroke=0)
        canv.setFillColor(ORANGE)
        canv.setFont("Helvetica-Bold", 10)
        canv.drawString(LM, PH - 7*mm, "SEOTool")
        logo_w = canv.stringWidth("SEOTool", "Helvetica-Bold", 10)
        canv.setFillColor(WHITE)
        canv.setFont("Helvetica", 8)
        canv.drawString(LM + logo_w + 8, PH - 7*mm, "Full SEO Audit Report")
        # URL right-aligned in header
        canv.setFont("Helvetica", 7)
        canv.drawRightString(PW - RM, PH - 7*mm, url[:70])
        # Footer
        canv.setFillColor(GRAY_LITE)
        canv.rect(0, 0, PW, 8*mm, fill=1, stroke=0)
        canv.setFillColor(GRAY)
        canv.setFont("Helvetica", 7)
        canv.drawString(LM, 10, generated)
        canv.drawRightString(PW - RM, 10, f"Page {doc.page}")
        canv.restoreState()

    # ══════════════════════════════════════════════
    #  DATA UNPACKING
    # ══════════════════════════════════════════════
    url        = data.get("url", "")
    audit_data = data.get("auditData", {})
    speed_data = data.get("speedData", {})
    tech_data  = data.get("techData", {})
    score_data = data.get("scoreData", {})
    crawl_data = data.get("crawlData", {})
    backlink   = data.get("backlinkData", {})
    mobile     = data.get("mobileData", {})
    generated  = f"Generated: {_now()}"

    buf = io.BytesIO()

    import functools
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=LM, rightMargin=RM,
        topMargin=TM + 10*mm,   # extra for top bar
        bottomMargin=BM + 8*mm  # extra for footer
    )

    story = []

    # ══════════════════════════════════════════════
    #  PAGE 1 — COVER
    # ══════════════════════════════════════════════
    story.append(Spacer(1, 28*mm))

    # Big brand name
    story.append(Paragraph("SEOTool", sty("logo2", fontSize=44, fontName="Helvetica-Bold",
                                          textColor=ORANGE, spaceAfter=4, alignment=TA_CENTER)))
    story.append(Paragraph("Full SEO Audit Report",
                            sty("rtype", fontSize=16, fontName="Helvetica",
                                textColor=DARK2, spaceAfter=6, alignment=TA_CENTER)))

    story.append(Spacer(1, 6*mm))

    # Orange divider line
    story.append(HRFlowable(width="60%", thickness=2, color=ORANGE,
                             hAlign="CENTER", spaceBefore=0, spaceAfter=6*mm))

    # URL box
    url_table = Table(
        [[cell(url, bold=True, color=DARK2, size=10, align=TA_CENTER)]],
        colWidths=[W * 0.85]
    )
    url_table.setStyle(TableStyle([
        ("BACKGROUND",    (0,0), (-1,-1), ORANGE_LITE),
        ("ROUNDEDCORNERS",(0,0), (-1,-1), [6,6,6,6]),
        ("TOPPADDING",    (0,0), (-1,-1), 10),
        ("BOTTOMPADDING", (0,0), (-1,-1), 10),
        ("LEFTPADDING",   (0,0), (-1,-1), 16),
        ("RIGHTPADDING",  (0,0), (-1,-1), 16),
    ]))
    story.append(Table([[url_table]], colWidths=[W],
                        style=[("ALIGN",(0,0),(-1,-1),"CENTER")]))

    story.append(Spacer(1, 8*mm))

    # Score cards row on cover
    ps   = (score_data.get("page_score") or score_data.get("score")) if score_data else None
    perf = speed_data.get("performance_score") if speed_data else None
    acc  = speed_data.get("accessibility_score") if speed_data else None
    bp   = speed_data.get("best_practices_score") if speed_data else None
    seo_psi = speed_data.get("seo_score") if speed_data else None

    gauge_items = [
        (ps,      "Overall SEO"),
        (perf,    "Performance"),
        (acc,     "Accessibility"),
        (bp,      "Best Practices"),
        (seo_psi, "SEO (PSI)"),
    ]

    # Score summary — circular gauges (uses the ScoreGauge flowable defined above)
    gauge_col_w = W / 5
    gauge_cells = []
    for score, label in gauge_items:
        gauge = ScoreGauge(score if score is not None else 0, label, size=46)
        wrapper = Table([[gauge]], colWidths=[gauge_col_w])
        wrapper.setStyle(TableStyle([
            ("ALIGN",  (0,0), (-1,-1), "CENTER"),
            ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ]))
        gauge_cells.append(wrapper)

    gauge_table = Table([gauge_cells], colWidths=[gauge_col_w] * 5,
                         hAlign="CENTER")
    gauge_table.setStyle(TableStyle([
        ("ALIGN",        (0,0), (-1,-1), "CENTER"),
        ("VALIGN",       (0,0), (-1,-1), "MIDDLE"),
        ("LEFTPADDING",  (0,0), (-1,-1), 2),
        ("RIGHTPADDING", (0,0), (-1,-1), 2),
        ("TOPPADDING",   (0,0), (-1,-1), 6),
        ("BOTTOMPADDING",(0,0), (-1,-1), 6),
    ]))
    story.append(gauge_table)

    # ══════════════════════════════════════════════
    #  AT A GLANCE — quick stats + verdict + top priorities
    #  (fills out the cover page with an instant, no-reading-required summary)
    # ══════════════════════════════════════════════
    story.append(Spacer(1, 10*mm))
    story.append(hr(thick=0.75, space_before=0, space_after=10))

    links_q = audit_data.get("links", {}) if audit_data else {}
    mobile_summary = mobile.get("summary", {}) if mobile else {}

    https_ok   = bool(tech_data.get("https")) if tech_data else False
    mobile_ok  = bool(tech_data.get("mobile_friendly") or mobile_summary.get("mobile_friendly")) if (tech_data or mobile) else False
    pages_crawled = crawl_data.get("total", 0) if crawl_data else 0
    broken_links  = links_q.get("broken_count", crawl_data.get("broken", 0) if crawl_data else 0)
    ref_domains   = backlink.get("unique_external_domains", 0) if backlink else 0
    schema_ok     = bool(tech_data.get("has_schema")) if tech_data else False

    stat_items = [
        ("HTTPS Secure",      "Yes" if https_ok else "No",         https_ok),
        ("Mobile Friendly",   "Yes" if mobile_ok else "No",        mobile_ok),
        ("Pages Crawled",     str(pages_crawled),                  True),
        ("Broken Links",      str(broken_links),                   broken_links == 0),
        ("Referring Domains", str(ref_domains),                    True),
        ("Structured Data",   "Yes" if schema_ok else "No",        schema_ok),
    ]

    story.append(Paragraph("At a Glance", H2))
    story.append(Spacer(1, 3))

    stat_col_w = W / 3
    stat_cells = []
    for label, value, good in stat_items:
        col = GREEN if good else RED
        bg  = GREEN_BG if good else RED_BG
        box = Table(
            [[cell(value, bold=True, color=col, size=14, align=TA_CENTER)],
             [cell(label, color=GRAY, size=7.5, align=TA_CENTER)]],
            colWidths=[stat_col_w - 4*mm]
        )
        box.setStyle(TableStyle([
            ("BACKGROUND",    (0,0), (-1,-1), bg),
            ("TOPPADDING",    (0,0), (-1,-1), 7),
            ("BOTTOMPADDING", (0,0), (-1,-1), 7),
        ]))
        stat_cells.append(box)

    # 3 columns x 2 rows
    stat_rows = [stat_cells[0:3], stat_cells[3:6]]
    stat_table = Table(stat_rows, colWidths=[stat_col_w]*3, hAlign="CENTER")
    stat_table.setStyle(TableStyle([
        ("ALIGN",        (0,0), (-1,-1), "CENTER"),
        ("VALIGN",       (0,0), (-1,-1), "MIDDLE"),
        ("LEFTPADDING",  (0,0), (-1,-1), 2*mm),
        ("RIGHTPADDING", (0,0), (-1,-1), 2*mm),
        ("TOPPADDING",   (0,0), (-1,-1), 2*mm),
        ("BOTTOMPADDING",(0,0), (-1,-1), 2*mm),
    ]))
    story.append(stat_table)

    story.append(Spacer(1, 6*mm))

    # ── Plain-language verdict, auto-written from the overall score ──
    if ps is not None:
        if ps >= 80:
            verdict = (f"<b>Overall, this site is in good shape</b> with a score of {int(ps)}/100. "
                       f"Most fundamentals are solid — focus on the smaller items below to push even higher.")
        elif ps >= 50:
            verdict = (f"<b>This site needs some work</b> — it scored {int(ps)}/100. "
                       f"There are clear opportunities to improve rankings and performance, detailed in the pages ahead.")
        else:
            verdict = (f"<b>This site needs significant improvement</b>, scoring only {int(ps)}/100. "
                       f"Several important issues were found that are likely hurting search visibility.")
        story.append(Paragraph(verdict, sty("verdict", fontSize=9.5, textColor=DARK2,
                                             alignment=TA_CENTER, leading=14)))
        story.append(Spacer(1, 6*mm))

    # ── Top priorities — short, plain-language punch list ──
    priorities = []
    if not https_ok:
        priorities.append("Switch the site to HTTPS — browsers flag non-secure sites, hurting trust and rankings.")
    if broken_links and broken_links > 0:
        priorities.append(f"Fix {broken_links} broken link(s) — these hurt user experience and crawlability.")
    if not mobile_ok:
        priorities.append("Improve mobile-friendliness — most search traffic is mobile-first.")
    if not schema_ok:
        priorities.append("Add structured data (schema) — helps search engines understand the content.")
    if perf is not None and perf < 50:
        priorities.append("Page speed is slow — optimize images/scripts to improve performance score.")

    if priorities:
        story.append(Paragraph("Top Priorities", H3))
        for p in priorities[:4]:
            row = Table([[cell("●", color=ORANGE, bold=True, size=9), cell(p, color=DARK, size=8.5)]],
                        colWidths=[6*mm, W - 6*mm])
            row.setStyle(TableStyle([
                ("VALIGN", (0,0), (-1,-1), "TOP"),
                ("LEFTPADDING", (0,0), (-1,-1), 0),
                ("TOPPADDING", (0,0), (-1,-1), 2),
                ("BOTTOMPADDING", (0,0), (-1,-1), 4),
            ]))
            story.append(row)
    else:
        story.append(Paragraph("✅ No major issues detected — nice work!",
                                sty("noissue", fontSize=9, textColor=GREEN, alignment=TA_CENTER)))

    story.append(Spacer(1, 6*mm))
    story.append(Paragraph(generated,
                            sty("gd", fontSize=8.5, textColor=GRAY, alignment=TA_CENTER)))

    story.append(PageBreak())

    # ══════════════════════════════════════════════
    #  SECTION: PageSpeed & Core Web Vitals
    # ══════════════════════════════════════════════
    if speed_data:
        story.append(SectionHeader("PageSpeed & Core Web Vitals", W, ""))
        story.append(Spacer(1, 5))

        def rating_pill(r):
            r = (r or "").strip()
            if not r: return cell("")
            return cell(r, bold=True, color=rating_color(r))

        vitals_rows = [
            [cell("Metric", bold=True), cell("Value", bold=True), cell("Rating", bold=True)],
            [cell("First Contentful Paint"),   cell(_ms_to_s(speed_data.get("fcp_ms"))),  rating_pill(speed_data.get("fcp_rating",""))],
            [cell("Largest Contentful Paint"), cell(_ms_to_s(speed_data.get("lcp_ms"))),  rating_pill(speed_data.get("lcp_rating",""))],
            [cell("Total Blocking Time"),      cell(_ms_to_s(speed_data.get("tbt_ms"))),  rating_pill(speed_data.get("tbt_rating",""))],
            [cell("Cumulative Layout Shift"),  cell(str(speed_data.get("cls","N/A"))),     rating_pill(speed_data.get("cls_rating",""))],
            [cell("Speed Index"),              cell(_ms_to_s(speed_data.get("si_ms"))),    cell("")],
            [cell("Time to Interactive"),      cell(_ms_to_s(speed_data.get("tti_ms"))),   cell("")],
            [cell("TTFB"),                     cell(_ms_to_s(speed_data.get("ttfb_ms"))),  rating_pill(speed_data.get("ttfb_rating",""))],
        ]
        story.append(make_table(vitals_rows, [0.47, 0.3, 0.23]))

        story.append(Spacer(1, 8))
        story.append(Paragraph("Resource Breakdown", H3))
        res_rows = [
            [cell("Resource", bold=True), cell("Size", bold=True)],
            [cell("JavaScript"), cell(f"{speed_data.get('total_js_kb',0)} KB")],
            [cell("CSS"),        cell(f"{speed_data.get('total_css_kb',0)} KB")],
            [cell("Images"),     cell(f"{speed_data.get('total_img_kb',0)} KB")],
            [cell("Fonts"),      cell(f"{speed_data.get('total_font_kb',0)} KB")],
            [cell("Total Page"), cell(f"{speed_data.get('page_size_kb',0)} KB")],
            [cell("Requests"),   cell(str(speed_data.get('total_requests',0)))],
        ]
        story.append(make_table(res_rows, [0.6, 0.4]))

        opps = speed_data.get("opportunities", [])
        if opps:
            story.append(Spacer(1, 8))
            story.append(Paragraph("Performance Opportunities", H3))
            opp_rows = [[cell("Opportunity", bold=True), cell("Saving", bold=True), cell("Details", bold=True)]]
            for o in opps:
                opp_rows.append([
                    cell(o.get("title", o.get("id", "")), color=DARK),
                    cell(o.get("displayValue", o.get("savings", "")), color=AMBER, bold=True),
                    cell(o.get("description", ""), color=GRAY, size=8),
                ])
            story.append(make_table(opp_rows, [0.35, 0.2, 0.45]))

    # ══════════════════════════════════════════════
    #  SECTION: On-Page SEO
    # ══════════════════════════════════════════════
    seo = audit_data.get("seo", {})
    if seo:
        story.append(Spacer(1, 10))
        story.append(SectionHeader("On-Page SEO", W, ""))
        story.append(Spacer(1, 5))

        title_d = seo.get("title", {})
        meta_d  = seo.get("meta_description", {})
        hdg     = seo.get("headings", {})
        cont    = seo.get("content", {})
        imgs    = seo.get("images", {})

        # Meta preview box
        meta_box_rows = [
            [cell("Title Tag", bold=True, color=ORANGE),
             cell(title_d.get("text", "—") or "—", color=DARK2)],
            [cell("Meta Description", bold=True, color=BLUE),
             cell(meta_d.get("text", "—") or "—", color=DARK)],
        ]
        mt = Table(meta_box_rows, colWidths=[W * 0.22, W * 0.78])
        mt.setStyle(TableStyle([
            ("BACKGROUND",    (0,0), (-1,-1), GRAY_LITE),
            ("TOPPADDING",    (0,0), (-1,-1), 7),
            ("BOTTOMPADDING", (0,0), (-1,-1), 7),
            ("LEFTPADDING",   (0,0), (-1,-1), 10),
            ("LINEBELOW",     (0,0), (-1,0), 0.5, BORDER),
        ]))
        story.append(mt)
        story.append(Spacer(1, 6))

        # On-page checks table
        def ok_fail(flag):
            return cell("Pass", bold=True, color=GREEN) if flag else cell("Fail", bold=True, color=RED)

        onpage_rows = [
            [cell("Check", bold=True), cell("Value", bold=True), cell("Status", bold=True)],
            [cell("Title Length"),     cell(f"{title_d.get('length',0)} chars"),   ok_fail(title_d.get("length_ok"))],
            [cell("Meta Present"),     cell("Yes" if meta_d.get("present") else "No"), ok_fail(meta_d.get("present"))],
            [cell("Meta Length"),      cell(f"{meta_d.get('length',0)} chars"),    ok_fail(meta_d.get("length_ok"))],
            [cell("H1 Tag"),           cell(hdg.get("h1_texts",[""])[0] if hdg.get("h1_texts") else "Missing"),
                                       ok_fail(hdg.get("h1_present"))],
            [cell("H1 Count"),         cell(str(hdg.get("h1_count",0))),           ok_fail(hdg.get("single_h1"))],
            [cell("H2 Count"),         cell(str(hdg.get("h2_count",0))),           cell("")],
            [cell("Word Count"),       cell(str(cont.get("word_count",0))),        ok_fail(cont.get("word_count_ok"))],
            [cell("Readability"),      cell(cont.get("readability_label","")),     cell("")],
            [cell("Images Total"),     cell(str(imgs.get("total",0))),             cell("")],
            [cell("Missing Alt Tags"), cell(str(imgs.get("missing_alt_count",0))),
                                       ok_fail(imgs.get("missing_alt_count",0)==0)],
            [cell("Broken Images"),    cell(str(imgs.get("broken_count",0))),
                                       ok_fail(imgs.get("broken_count",0)==0)],
        ]
        story.append(make_table(onpage_rows, [0.38, 0.42, 0.2]))

        missing_alt_imgs = imgs.get("missing_alt", [])
        if missing_alt_imgs:
            story.append(Spacer(1, 6))
            story.append(Paragraph(f"Images Missing Alt Text  ({len(missing_alt_imgs)} found)", H3))
            img_rows = [[cell("#", bold=True), cell("Image src URL", bold=True)]]
            for idx, img in enumerate(missing_alt_imgs, 1):
                src = img.get("src", img.get("url","")) if isinstance(img, dict) else str(img)
                img_rows.append([cell(str(idx), color=ORANGE, bold=True), cell(src, color=GRAY, size=8)])
            story.append(make_table(img_rows, [0.06, 0.94]))

        kw = seo.get("keywords", {})
        if kw:
            kw_list = kw.get("top_keywords", kw.get("keywords", []))
            if kw_list:
                story.append(Spacer(1,6))
                story.append(Paragraph("Top Keywords", H3))
                kw_rows = [[cell("Keyword", bold=True), cell("Count / Density", bold=True)]]
                for k in kw_list[:20]:
                    if isinstance(k, dict):
                        kw_rows.append([cell(k.get("word", k.get("keyword",""))),
                                        cell(str(k.get("count", k.get("density",""))))])
                    else:
                        kw_rows.append([cell(str(k)), cell("")])
                story.append(make_table(kw_rows, [0.7, 0.3]))

    # ══════════════════════════════════════════════
    #  SECTION: Link Analysis
    # ══════════════════════════════════════════════
    links = audit_data.get("links", {})
    if links:
        story.append(Spacer(1, 10))
        story.append(SectionHeader("Link Analysis", W, ""))
        story.append(Spacer(1, 5))

        link_rows = [
            [cell("Metric", bold=True), cell("Count", bold=True)],
            [cell("Internal Links"),  cell(str(links.get("internal_count",0)))],
            [cell("External Links"),  cell(str(links.get("external_count",0)))],
            [cell("Broken Links"),    cell(str(links.get("broken_count",0)),
                                          color=RED if links.get("broken_count",0) > 0 else GREEN,
                                          bold=links.get("broken_count",0) > 0)],
            [cell("Nofollow Links"),  cell(str(links.get("nofollow_count",0)))],
            [cell("Total Links"),     cell(str(links.get("total_links",0)))],
        ]
        story.append(make_table(link_rows, [0.6, 0.4]))

        broken = links.get("broken_links", [])
        if broken:
            story.append(Spacer(1, 6))
            story.append(Paragraph(f"Broken Links  ({len(broken)} found)", H3))
            b_rows = [[cell("URL", bold=True), cell("Status", bold=True), cell("Anchor Text", bold=True)]]
            for b in broken:
                b_rows.append([
                    cell(b.get("url",""), color=GRAY, size=8),
                    cell(str(b.get("status","")), color=RED, bold=True),
                    cell(b.get("anchor",""), color=DARK),
                ])
            story.append(make_table(b_rows, [0.55, 0.12, 0.33]))

        anchors = links.get("top_anchors", links.get("anchors", []))
        if anchors:
            story.append(Spacer(1, 6))
            story.append(Paragraph("Top Anchor Texts", H3))
            a_rows = [[cell("Anchor Text", bold=True), cell("Count", bold=True)]]
            for a in anchors[:20]:
                if isinstance(a, dict):
                    a_rows.append([cell(a.get("text", a.get("anchor",""))),
                                   cell(str(a.get("count","")))])
                else:
                    a_rows.append([cell(str(a)), cell("")])
            story.append(make_table(a_rows, [0.75, 0.25]))

    # ══════════════════════════════════════════════
    #  SECTION: Technical SEO
    # ══════════════════════════════════════════════
    if tech_data:
        story.append(Spacer(1, 10))
        story.append(SectionHeader("Technical SEO", W, ""))
        story.append(Spacer(1, 5))

        tech_items = [
            ("HTTPS",           "Yes" if tech_data.get("https") else "No", ""),
            ("robots.txt",      "Found" if tech_data.get("robots_txt") else "Missing", ""),
            ("sitemap.xml",     "Found" if tech_data.get("sitemap_xml") else "Missing",
                                f"{tech_data.get('sitemap_url_count',0)} URLs"),
            ("Canonical Tag",   "Present" if tech_data.get("canonical") else "Missing",
                                tech_data.get("canonical_url","")),
            ("Structured Data", "Yes" if tech_data.get("has_schema") else "No", ""),
            ("Open Graph",      "Yes" if tech_data.get("has_og") else "No", ""),
            ("Mobile Friendly", "Yes" if tech_data.get("mobile_friendly") else "No", ""),
            ("Noindex",         "Yes" if tech_data.get("noindex") else "No", ""),
            ("Load Time",       _ms_to_s(tech_data.get("load_time_ms")), ""),
            ("Score",           str(tech_data.get("score","")), ""),
        ]
        t_rows = [[cell("Check", bold=True), cell("Status", bold=True), cell("Detail", bold=True)]]
        for check, status, detail in tech_items:
            t_rows.append(status_row(check, status, detail))
        story.append(make_table(t_rows, [0.35, 0.28, 0.37]))

        schema_types = tech_data.get("schema_types", tech_data.get("structured_data_types", []))
        if schema_types:
            story.append(Spacer(1, 6))
            story.append(Paragraph("Structured Data / Schema Types", H3))
            st_list = schema_types if isinstance(schema_types, list) else [schema_types]
            st_rows = [[cell("#", bold=True), cell("Schema Type", bold=True)]]
            for i, stype in enumerate(st_list, 1):
                st_rows.append([cell(str(i), color=ORANGE, bold=True), cell(str(stype))])
            story.append(make_table(st_rows, [0.08, 0.92]))

        og_data = tech_data.get("og_data", tech_data.get("open_graph", {}))
        if og_data and isinstance(og_data, dict) and og_data:
            story.append(Spacer(1, 6))
            story.append(Paragraph("Open Graph Tags", H3))
            og_rows = [[cell("Property", bold=True), cell("Content", bold=True)]]
            for k, v in og_data.items():
                og_rows.append([cell(str(k), bold=True, color=BLUE), cell(str(v), color=DARK)])
            story.append(make_table(og_rows, [0.3, 0.7]))

    # ══════════════════════════════════════════════
    #  SECTION: Backlink Analysis
    # ══════════════════════════════════════════════
    if backlink:
        story.append(Spacer(1, 10))
        story.append(SectionHeader("Backlink Analysis", W, ""))
        story.append(Spacer(1, 5))

        bl_rows = [
            [cell("Metric", bold=True), cell("Value", bold=True)],
            [cell("External Links"),         cell(str(backlink.get("external_links",0)))],
            [cell("Unique Domains"),         cell(str(backlink.get("unique_external_domains",0)))],
            [cell("Dofollow"),               cell(str(backlink.get("dofollow_count",0)), color=GREEN, bold=True)],
            [cell("Nofollow"),               cell(str(backlink.get("nofollow_count",0)))],
            [cell("High Authority Links"),   cell(str(backlink.get("high_authority_links",0)), color=GREEN, bold=True)],
            [cell("Broken External"),        cell(str(backlink.get("broken_external_count",0)),
                                                  color=RED if backlink.get("broken_external_count",0) else GREEN,
                                                  bold=True)],
            [cell("Link Equity Score"),      cell(str(backlink.get("link_equity_score",0)))],
            [cell("Overall Score"),          cell(str(backlink.get("overall_score",0)),
                                                  color=score_color(backlink.get("overall_score")), bold=True)],
        ]
        story.append(make_table(bl_rows, [0.6, 0.4]))

        top_domains = backlink.get("top_domains", backlink.get("top_referring_domains", []))
        if top_domains:
            story.append(Spacer(1, 6))
            story.append(Paragraph("Top Referring Domains", H3))
            td_rows = [[cell("Domain", bold=True), cell("Links", bold=True)]]
            for d in top_domains[:15]:
                if isinstance(d, dict):
                    td_rows.append([cell(d.get("domain","")), cell(str(d.get("count","")))])
                else:
                    td_rows.append([cell(str(d)), cell("")])
            story.append(make_table(td_rows, [0.75, 0.25]))

    # ══════════════════════════════════════════════
    #  SECTION: Crawler Results
    # ══════════════════════════════════════════════
    if crawl_data and crawl_data.get("total"):
        story.append(Spacer(1, 10))
        story.append(SectionHeader("Site Crawler Results", W, ""))
        story.append(Spacer(1, 5))

        cr_rows = [
            [cell("Metric", bold=True), cell("Count", bold=True)],
            [cell("Total Pages"),  cell(str(crawl_data.get("total",0)))],
            [cell("OK (200)"),     cell(str(crawl_data.get("ok",0)), color=GREEN, bold=True)],
            [cell("Broken"),       cell(str(crawl_data.get("broken",0)),
                                        color=RED if crawl_data.get("broken",0) else GREEN, bold=True)],
            [cell("Redirects"),    cell(str(crawl_data.get("redirects",0)), color=AMBER)],
            [cell("Errors"),       cell(str(crawl_data.get("errors",0)),
                                        color=RED if crawl_data.get("errors",0) else GREEN)],
            [cell("Orphan Pages"), cell(str(crawl_data.get("orphans",0)))],
            [cell("Crawl Time"),   cell(f"{crawl_data.get('elapsed',0)}s")],
        ]
        story.append(make_table(cr_rows, [0.6, 0.4]))

        pages = crawl_data.get("pages", [])
        if pages:
            story.append(Spacer(1, 8))
            story.append(Paragraph("Crawled Pages Detail", H3))
            pg_rows = [[cell("URL", bold=True), cell("Status", bold=True),
                        cell("SEO Score", bold=True), cell("Depth", bold=True), cell("Orphan", bold=True)]]
            for p in pages[:50]:
                seo_s = str(p.get("seo",{}).get("seo_score","—")) if p.get("seo") else "—"
                st = str(p.get("status",""))
                pg_rows.append([
                    cell(p.get("url",""), color=GRAY, size=7.5),
                    cell(st, color=GREEN if st.startswith("2") else RED, bold=True),
                    cell(seo_s),
                    cell(str(p.get("depth",0))),
                    cell("Yes" if p.get("orphan") else "No",
                         color=AMBER if p.get("orphan") else DARK),
                ])
            story.append(make_table(pg_rows, [0.49, 0.12, 0.14, 0.1, 0.15], compact=True))

    # ══════════════════════════════════════════════
    #  SECTION: Mobile SEO
    # ══════════════════════════════════════════════
    if mobile:
        summary_m = mobile.get("summary", {})
        story.append(Spacer(1, 10))
        story.append(SectionHeader("Mobile SEO", W, ""))
        story.append(Spacer(1, 5))

        mob_score = summary_m.get("score", 0)
        mob_rows = [
            [cell("Check", bold=True), cell("Value", bold=True)],
            [cell("Mobile Score"),    cell(str(mob_score),
                                           color=score_color(mob_score), bold=True)],
            [cell("Mobile Friendly"), cell("Yes" if summary_m.get("mobile_friendly") else "No",
                                           color=GREEN if summary_m.get("mobile_friendly") else RED,
                                           bold=True)],
            [cell("Passed Checks"),   cell(f"{summary_m.get('passed',0)} / {summary_m.get('total',0)}")],
        ]
        cwv = mobile.get("core_web_vitals", {})
        if cwv:
            mob_rows += [
                [cell("LCP"),  cell(_ms_to_s(cwv.get("lcp_ms")),
                                    color=rating_color(cwv.get("lcp_rating","")))],
                [cell("CLS"),  cell(str(cwv.get("cls","")))],
                [cell("INP"),  cell(_ms_to_s(cwv.get("inp_ms")))],
            ]
        story.append(make_table(mob_rows, [0.6, 0.4]))

    # ══════════════════════════════════════════════
    #  FOOTER NOTE
    # ══════════════════════════════════════════════
    story.append(Spacer(1, 12))
    story.append(hr(space_before=4, space_after=6))
    story.append(Paragraph(
        f"Report generated by <b>SEOBit</b> &nbsp;·&nbsp; {_now()}",
        sty("foot", fontSize=8, textColor=GRAY, alignment=TA_CENTER)
    ))

    # ══════════════════════════════════════════════
    #  BUILD with page template
    # ══════════════════════════════════════════════
    doc.build(
        story,
        onFirstPage=functools.partial(on_page, url=url, generated=generated),
        onLaterPages=functools.partial(on_page, url=url, generated=generated),
    )
    return buf.getvalue()

# ─────────────────────────────────────────────────────────────────────────────
#  EXCEL EXPORT
# ─────────────────────────────────────────────────────────────────────────────

def generate_excel(data: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, Fill, PatternFill, Alignment, Border, Side, GradientFill
    )
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    url        = data.get("url", "")
    audit_data = data.get("auditData", {})
    speed_data = data.get("speedData", {})
    tech_data  = data.get("techData", {})
    score_data = data.get("scoreData", {})
    crawl_data = data.get("crawlData", {})
    backlink   = data.get("backlinkData", {})
    mobile     = data.get("mobileData", {})

    # ── Style helpers ──
    HDR_FILL  = PatternFill("solid", fgColor="F06400")
    ALT_FILL  = PatternFill("solid", fgColor="F1F0EC")
    WHITE_FILL= PatternFill("solid", fgColor="FFFFFF")
    HDR_FONT  = Font(bold=True, color="FFFFFF", size=10)
    BOLD_FONT = Font(bold=True, size=10)
    NORM_FONT = Font(size=9)
    thin = Side(style="thin", color="D4D1C8")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    def add_sheet(name):
        ws = wb.create_sheet(title=name)
        return ws

    def hdr_row(ws, row, cols, row_num=1):
        for c, val in enumerate(cols, 1):
            cell = ws.cell(row=row_num, column=c, value=val)
            cell.fill = HDR_FILL
            cell.font = HDR_FONT
            cell.alignment = CENTER
            cell.border = BORDER
        ws.row_dimensions[row_num].height = 18

    def data_row(ws, row_num, cols, alt=False):
        fill = ALT_FILL if alt else WHITE_FILL
        for c, val in enumerate(cols, 1):
            cell = ws.cell(row=row_num, column=c, value=val)
            cell.fill = fill
            cell.font = NORM_FONT
            cell.alignment = LEFT
            cell.border = BORDER

    def set_col_widths(ws, widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def section_title(ws, row_num, text):
        cell = ws.cell(row=row_num, column=1, value=text)
        cell.font = Font(bold=True, size=10, color="F06400")
        return row_num + 1

    # ════════════════════════════════
    #  SHEET 1 — Summary
    # ════════════════════════════════
    ws = add_sheet("Summary")
    ws.merge_cells("A1:D1")
    title_cell = ws["A1"]
    title_cell.value = f"SEOBit Audit — {url}"
    title_cell.font  = Font(bold=True, size=14, color="F06400")
    title_cell.alignment = LEFT
    ws.row_dimensions[1].height = 22

    ws["A2"].value = f"Generated: {_now()}"
    ws["A2"].font  = Font(size=9, color="7a7870")

    hdr_row(ws, None, ["Section", "Score", "Status", "Notes"], row_num=4)
    rows = []

    ps = score_data.get("page_score") or score_data.get("score")
    rows.append(["Overall SEO Score", ps, _score_label(ps), ""])

    if speed_data:
        rows.append(["Performance",    speed_data.get("performance_score"),   _score_label(speed_data.get("performance_score")),  ""])
        rows.append(["Accessibility",  speed_data.get("accessibility_score"), _score_label(speed_data.get("accessibility_score")), ""])
        rows.append(["Best Practices", speed_data.get("best_practices_score"),_score_label(speed_data.get("best_practices_score")), ""])
        rows.append(["SEO (PSI)",      speed_data.get("seo_score"),           _score_label(speed_data.get("seo_score")),           ""])

    if tech_data:
        rows.append(["Technical SEO Score", tech_data.get("score"), _score_label(tech_data.get("score")), ""])

    if mobile:
        m_score = mobile.get("summary",{}).get("score")
        rows.append(["Mobile SEO Score", m_score, _score_label(m_score), ""])

    if backlink:
        rows.append(["Backlink Score", backlink.get("overall_score"), _score_label(backlink.get("overall_score")), ""])

    for i, row in enumerate(rows):
        data_row(ws, i+5, row, alt=i%2==1)

    set_col_widths(ws, [28, 12, 16, 30])

    # ════════════════════════════════
    #  SHEET 2 — PageSpeed
    # ════════════════════════════════
    if speed_data:
        ws2 = add_sheet("PageSpeed")
        hdr_row(ws2, None, ["Metric", "Value", "Rating"], row_num=1)
        vitals = [
            ("First Contentful Paint",   _ms_to_s(speed_data.get("fcp_ms")),  speed_data.get("fcp_rating","")),
            ("Largest Contentful Paint", _ms_to_s(speed_data.get("lcp_ms")),  speed_data.get("lcp_rating","")),
            ("Total Blocking Time",      _ms_to_s(speed_data.get("tbt_ms")),  speed_data.get("tbt_rating","")),
            ("Cumulative Layout Shift",  speed_data.get("cls",""),             speed_data.get("cls_rating","")),
            ("Speed Index",              _ms_to_s(speed_data.get("si_ms")),   ""),
            ("Time to Interactive",      _ms_to_s(speed_data.get("tti_ms")),  ""),
            ("TTFB",                     _ms_to_s(speed_data.get("ttfb_ms")), speed_data.get("ttfb_rating","")),
            ("","",""),
            ("JavaScript",               f"{speed_data.get('total_js_kb',0)} KB",  ""),
            ("CSS",                      f"{speed_data.get('total_css_kb',0)} KB",  ""),
            ("Images",                   f"{speed_data.get('total_img_kb',0)} KB",  ""),
            ("Fonts",                    f"{speed_data.get('total_font_kb',0)} KB",  ""),
            ("Total Page Size",          f"{speed_data.get('page_size_kb',0)} KB",   ""),
            ("Total Requests",           speed_data.get("total_requests",0),          ""),
            ("Unused JS",                f"{speed_data.get('unused_js_kb',0)} KB",   ""),
            ("Unused CSS",               f"{speed_data.get('unused_css_kb',0)} KB",  ""),
        ]
        for i, row in enumerate(vitals):
            data_row(ws2, i+2, list(row), alt=i%2==1)
        set_col_widths(ws2, [28, 18, 18])

        # ── Opportunities sub-section ──
        opps = speed_data.get("opportunities", [])
        if opps:
            opp_start = len(vitals) + 4
            section_title(ws2, opp_start, "Performance Opportunities")
            hdr_row(ws2, None, ["Opportunity", "Potential Saving", "Description"], row_num=opp_start+1)
            for i, o in enumerate(opps):
                data_row(ws2, opp_start+2+i, [
                    o.get("title", o.get("id", "")),
                    o.get("displayValue", o.get("savings", "")),
                    o.get("description", ""),
                ], alt=i%2==1)
            set_col_widths(ws2, [28, 18, 60])

    # ════════════════════════════════
    #  SHEET 3 — On-Page SEO
    # ════════════════════════════════
    seo = audit_data.get("seo", {})
    if seo:
        ws3 = add_sheet("On-Page SEO")
        hdr_row(ws3, None, ["Check", "Value", "Status"], row_num=1)
        title_d = seo.get("title", {})
        meta_d  = seo.get("meta_description", {})
        hdg     = seo.get("headings", {})
        cont    = seo.get("content", {})
        imgs    = seo.get("images", {})
        rows3 = [
            # Full text — no truncation
            ("Title Text",           title_d.get("text",""),           ""),
            ("Title Length",         title_d.get("length",0),          "✅" if title_d.get("length_ok") else "❌"),
            ("Meta Description",     meta_d.get("text",""),            "✅" if meta_d.get("present") else "❌"),
            ("Meta Length",          meta_d.get("length",0),           "✅" if meta_d.get("length_ok") else "❌"),
            ("H1",                   hdg.get("h1_texts",[""])[0] if hdg.get("h1_texts") else "Missing", "✅" if hdg.get("h1_present") else "❌"),
            ("H1 Count",             hdg.get("h1_count",0),            "✅" if hdg.get("single_h1") else "⚠️"),
            ("H2 Count",             hdg.get("h2_count",0),            ""),
            ("Word Count",           cont.get("word_count",0),         "✅" if cont.get("word_count_ok") else "❌"),
            ("Readability",          cont.get("readability_label",""), ""),
            ("Images Total",         imgs.get("total",0),              ""),
            ("Missing Alt Tags",     imgs.get("missing_alt_count",0),  "✅" if imgs.get("missing_alt_count",0)==0 else "❌"),
            ("Broken Images",        imgs.get("broken_count",0),       "✅" if imgs.get("broken_count",0)==0 else "❌"),
        ]
        for i, row in enumerate(rows3):
            data_row(ws3, i+2, list(row), alt=i%2==1)
        set_col_widths(ws3, [24, 70, 12])

    # ════════════════════════════════
    #  SHEET 4 — Images Missing Alt
    # ════════════════════════════════
    if seo:
        imgs = seo.get("images", {})
        missing_alt_imgs = imgs.get("missing_alt", [])
        if missing_alt_imgs:
            ws_img = add_sheet("Images Missing Alt")
            hdr_row(ws_img, None, ["#", "Image src URL", "Current Alt"], row_num=1)
            for i, img in enumerate(missing_alt_imgs, 1):
                if isinstance(img, dict):
                    src = img.get("src", img.get("url", ""))
                    alt_val = img.get("alt", "")
                else:
                    src = str(img)
                    alt_val = ""
                data_row(ws_img, i+1, [i, src, alt_val or "MISSING"], alt=i%2==1)
            set_col_widths(ws_img, [6, 80, 18])

    # ════════════════════════════════
    #  SHEET 5 — Links
    # ════════════════════════════════
    links = audit_data.get("links", {})
    if links:
        ws4 = add_sheet("Links")
        hdr_row(ws4, None, ["Metric", "Value"], row_num=1)
        link_summary = [
            ("Internal Links",   links.get("internal_count",0)),
            ("External Links",   links.get("external_count",0)),
            ("Broken Links",     links.get("broken_count",0)),
            ("Nofollow Links",   links.get("nofollow_count",0)),
            ("Total Links",      links.get("total_links",0)),
        ]
        for i, row in enumerate(link_summary):
            data_row(ws4, i+2, list(row), alt=i%2==1)

        # ── Broken links — every link, full URL + status + anchor ──
        broken = links.get("broken_links", [])
        if broken:
            brk_start = len(link_summary) + 4
            section_title(ws4, brk_start, f"Broken Links Detail ({len(broken)} found)")
            hdr_row(ws4, None, ["URL", "HTTP Status", "Anchor Text"], row_num=brk_start+1)
            for i, b in enumerate(broken):
                data_row(ws4, brk_start+2+i, [
                    b.get("url",""),
                    b.get("status",""),
                    b.get("anchor",""),
                ], alt=i%2==1)
            anchor_offset = brk_start + len(broken) + 4
        else:
            anchor_offset = len(link_summary) + 5

        # ── Top anchors ──
        anchors = links.get("top_anchors", links.get("anchors", []))
        if anchors:
            section_title(ws4, anchor_offset, "Top Anchor Texts")
            hdr_row(ws4, None, ["Anchor Text", "Count"], row_num=anchor_offset+1)
            for i, a in enumerate(anchors[:30]):
                if isinstance(a, dict):
                    data_row(ws4, anchor_offset+2+i, [a.get("text", a.get("anchor","")), a.get("count","")], alt=i%2==1)
                else:
                    data_row(ws4, anchor_offset+2+i, [str(a), ""], alt=i%2==1)

        set_col_widths(ws4, [60, 14, 30])

    # ════════════════════════════════
    #  SHEET 6 — Keywords
    # ════════════════════════════════
    if seo:
        kw = seo.get("keywords", {})
        kw_list = kw.get("top_keywords", kw.get("keywords", [])) if kw else []
        if kw_list:
            ws_kw = add_sheet("Keywords")
            hdr_row(ws_kw, None, ["Keyword", "Count", "Density %"], row_num=1)
            for i, k in enumerate(kw_list):
                if isinstance(k, dict):
                    data_row(ws_kw, i+2, [
                        k.get("word", k.get("keyword", "")),
                        k.get("count", ""),
                        k.get("density", ""),
                    ], alt=i%2==1)
                else:
                    data_row(ws_kw, i+2, [str(k), "", ""], alt=i%2==1)
            set_col_widths(ws_kw, [30, 12, 14])

    # ════════════════════════════════
    #  SHEET 7 — Technical SEO
    # ════════════════════════════════
    if tech_data:
        ws5 = add_sheet("Technical SEO")
        hdr_row(ws5, None, ["Check", "Status", "Detail"], row_num=1)
        tech_rows = [
            ("HTTPS",           "✅ Yes" if tech_data.get("https") else "❌ No",           ""),
            ("robots.txt",      "✅ Found" if tech_data.get("robots_txt") else "❌ Missing", ""),
            ("sitemap.xml",     "✅ Found" if tech_data.get("sitemap_xml") else "❌ Missing", f"{tech_data.get('sitemap_url_count',0)} URLs"),
            ("Canonical",       "✅ Yes" if tech_data.get("canonical") else "❌ No",         tech_data.get("canonical_url","")),
            ("Structured Data", "✅ Yes" if tech_data.get("has_schema") else "❌ No",        ""),
            ("Open Graph",      "✅ Yes" if tech_data.get("has_og") else "❌ No",            ""),
            ("Mobile Friendly", "✅ Yes" if tech_data.get("mobile_friendly") else "❌ No",  ""),
            ("Noindex",         "⚠️ Yes" if tech_data.get("noindex") else "✅ No",           ""),
            ("Load Time",       _ms_to_s(tech_data.get("load_time_ms")),                    ""),
            ("Score",           tech_data.get("score",""),                                   ""),
        ]
        for i, row in enumerate(tech_rows):
            data_row(ws5, i+2, list(row), alt=i%2==1)

        # ── Schema types detail ──
        schema_types = tech_data.get("schema_types", tech_data.get("structured_data_types", []))
        if schema_types:
            sch_start = len(tech_rows) + 4
            section_title(ws5, sch_start, "Structured Data / Schema Types")
            hdr_row(ws5, None, ["#", "Schema Type"], row_num=sch_start+1)
            st_list = schema_types if isinstance(schema_types, list) else [schema_types]
            for i, stype in enumerate(st_list, 1):
                data_row(ws5, sch_start+1+i, [i, str(stype)], alt=i%2==1)
            og_offset = sch_start + len(st_list) + 4
        else:
            og_offset = len(tech_rows) + 4

        # ── Open Graph tags detail ──
        og_data = tech_data.get("og_data", tech_data.get("open_graph", {}))
        if og_data and isinstance(og_data, dict) and og_data:
            section_title(ws5, og_offset, "Open Graph Tags")
            hdr_row(ws5, None, ["Property", "Content"], row_num=og_offset+1)
            for i, (k, v) in enumerate(og_data.items()):
                data_row(ws5, og_offset+2+i, [str(k), str(v)], alt=i%2==1)

        set_col_widths(ws5, [24, 16, 50])

    # ════════════════════════════════
    #  SHEET 8 — Backlinks
    # ════════════════════════════════
    if backlink:
        ws6 = add_sheet("Backlinks")
        hdr_row(ws6, None, ["Metric", "Value"], row_num=1)
        bl_rows = [
            ("Overall Score",         backlink.get("overall_score",0)),
            ("External Links",        backlink.get("external_links",0)),
            ("Unique Domains",        backlink.get("unique_external_domains",0)),
            ("Dofollow",              backlink.get("dofollow_count",0)),
            ("Nofollow",              backlink.get("nofollow_count",0)),
            ("High Authority Links",  backlink.get("high_authority_links",0)),
            ("Broken External",       backlink.get("broken_external_count",0)),
            ("Link Equity Score",     backlink.get("link_equity_score",0)),
        ]
        for i, row in enumerate(bl_rows):
            data_row(ws6, i+2, list(row), alt=i%2==1)

        top_domains = backlink.get("top_domains", backlink.get("top_referring_domains", []))
        if top_domains:
            td_start = len(bl_rows) + 4
            section_title(ws6, td_start, "Top Referring Domains")
            hdr_row(ws6, None, ["Domain", "Links"], row_num=td_start+1)
            for i, d in enumerate(top_domains[:20]):
                if isinstance(d, dict):
                    data_row(ws6, td_start+2+i, [d.get("domain",""), d.get("count","")], alt=i%2==1)
                else:
                    data_row(ws6, td_start+2+i, [str(d), ""], alt=i%2==1)

        set_col_widths(ws6, [40, 16])

    # ════════════════════════════════
    #  SHEET 9 — Crawler Results
    # ════════════════════════════════
    if crawl_data and crawl_data.get("total"):
        ws7 = add_sheet("Crawler Results")
        hdr_row(ws7, None, ["Metric", "Value"], row_num=1)
        cr_summary = [
            ("Total Pages",   crawl_data.get("total",0)),
            ("OK (200)",      crawl_data.get("ok",0)),
            ("Broken",        crawl_data.get("broken",0)),
            ("Redirects",     crawl_data.get("redirects",0)),
            ("Errors",        crawl_data.get("errors",0)),
            ("Orphans",       crawl_data.get("orphans",0)),
            ("Crawl Time",    f"{crawl_data.get('elapsed',0)}s"),
        ]
        for i, row in enumerate(cr_summary):
            data_row(ws7, i+2, list(row), alt=i%2==1)

        pages = crawl_data.get("pages", [])
        if pages:
            ws7.cell(row=11, column=1, value="Pages Detail").font = Font(bold=True, size=10)
            hdr_row(ws7, None, ["URL", "Status", "Category", "SEO Score", "Depth", "Orphan"], row_num=12)
            for i, p in enumerate(pages[:200]):
                seo_s = p.get("seo",{}).get("seo_score","") if p.get("seo") else ""
                data_row(ws7, i+13, [
                    p.get("url",""),
                    p.get("status",""),
                    p.get("category",""),
                    seo_s,
                    p.get("depth",0),
                    "Yes" if p.get("orphan") else "No"
                ], alt=i%2==1)
        set_col_widths(ws7, [55, 10, 14, 12, 8, 10])

    # ════════════════════════════════
    #  SHEET 10 — Mobile SEO
    # ════════════════════════════════
    if mobile:
        ws8 = add_sheet("Mobile SEO")
        summary_m = mobile.get("summary", {})
        hdr_row(ws8, None, ["Check", "Status", "Detail"], row_num=1)
        mob_rows = [
            ("Score",          summary_m.get("score",""),                                  ""),
            ("Mobile Friendly","✅ Yes" if summary_m.get("mobile_friendly") else "❌ No", ""),
            ("Passed Checks",  f"{summary_m.get('passed',0)}/{summary_m.get('total',0)}", ""),
        ]
        cwv = mobile.get("core_web_vitals", {})
        if cwv:
            mob_rows += [
                ("LCP",   _ms_to_s(cwv.get("lcp_ms")),  ""),
                ("CLS",   cwv.get("cls",""),             ""),
                ("INP",   _ms_to_s(cwv.get("inp_ms")),  ""),
            ]
        for i, row in enumerate(mob_rows):
            data_row(ws8, i+2, list(row), alt=i%2==1)
        set_col_widths(ws8, [24, 16, 30])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()